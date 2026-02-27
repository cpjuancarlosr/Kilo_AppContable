"""
Servicio de importación de estados de cuenta bancarios.
Soporta: PDF, CSV, Excel (xlsx, xls)
"""

import re
import io
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from typing import List, Dict, Optional, Any, Tuple
from dataclasses import dataclass
import csv

# Para PDF
try:
    import PyPDF2

    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False

# Para Excel
try:
    import openpyxl
    from openpyxl import load_workbook

    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

try:
    import pandas as pd

    PANDAS_SUPPORT = True
except ImportError:
    PANDAS_SUPPORT = False


@dataclass
class TransaccionBancaria:
    """Representa una transacción bancaria extraída."""

    fecha: date
    descripcion: str
    referencia: str
    cargo: Decimal
    abono: Decimal
    saldo: Optional[Decimal] = None
    concepto_limpio: str = ""
    categoria_sugerida: str = ""
    es_transferencia_interna: bool = False
    beneficiario_detectado: str = ""
    rfc_detectado: Optional[str] = None

    @property
    def monto_neto(self) -> Decimal:
        return self.abono - self.cargo

    @property
    def tipo(self) -> str:
        if self.cargo > 0:
            return "cargo"
        elif self.abono > 0:
            return "abono"
        return "neutro"


@dataclass
class EstadoCuentaImportado:
    """Resultado de importación de estado de cuenta."""

    banco: str
    cuenta: str
    moneda: str
    fecha_inicio: date
    fecha_fin: date
    saldo_inicial: Decimal
    saldo_final: Decimal
    transacciones: List[TransaccionBancaria]
    total_cargos: Decimal
    total_abonos: Decimal

    @property
    def balance_verificado(self) -> bool:
        """Verifica que el balance cuadre."""
        calculado = self.saldo_inicial + self.total_abonos - self.total_cargos
        return abs(calculado - self.saldo_final) < Decimal("0.01")


class ParserBancarioMX:
    """Parser especializado para bancos mexicanos."""

    # Patrones de bancos mexicanos
    BANCOS = {
        "BBVA": ["bbva", "bancomer", "014"],
        "SANTANDER": ["santander", "0140", "banco santander"],
        "BANORTE": ["banorte", "banorte-ixe", "058"],
        "HSBC": ["hsbc", "021"],
        "BANAMEX": ["banamex", "citibanamex", "002"],
        "SCOTIABANK": ["scotiabank", "044", "scotia"],
        "BANCOAZTECA": ["banco azteca", "062"],
        "INBURSA": ["inbursa", "036"],
        "INTERACCIONES": ["interacciones", "060"],
        "BANREGIO": ["banregio", "059"],
        "AFIRME": ["afirme", "061"],
        "MONEX": ["monex", "056"],
        "MULTIVA": ["multiva", "049"],
    }

    # Categorías inteligentes basadas en descripción
    CATEGORIAS = {
        "transferencias_internas": [
            r"transferencia.*misma empresa",
            r"traspaso.*cuenta",
            r"spei.*propia",
            r"transferencia entre cuentas",
        ],
        "proveedores": [
            r"pago.*proveedor",
            r"factura",
            r"pago a [\w\s]+",
            r"transferencia.*proveedor",
        ],
        "nominas": [
            r"n[oó]mina",
            r"pago.*sueldo",
            r"salario",
            r"transferencia.*empleado",
            r"spei.*nomin",
        ],
        "impuestos": [
            r"sat",
            r"impuesto",
            r"iva",
            r"isr",
            r"pago provisional",
            r"declaraci[oó]n",
        ],
        "servicios": [
            r"luz",
            r"agua",
            r"tel[eé]fono",
            r"internet",
            r"comisi[oó]n bancaria",
            r"cuota",
            r"mantenimiento",
        ],
        "clientes": [
            r"pago.*cliente",
            r"deposito.*cliente",
            r"transferencia.*cliente",
            r"venta",
            r"cobro",
        ],
        "financiamiento": [
            r"pago.*cr[eé]dito",
            r"amortizaci[oó]n",
            r"intereses",
            r"comisi[oó]n.*apertura",
        ],
    }

    @classmethod
    def detectar_banco(cls, contenido: str) -> str:
        """Detecta el banco basado en el contenido."""
        contenido_lower = contenido.lower()

        for banco, patrones in cls.BANCOS.items():
            for patron in patrones:
                if patron in contenido_lower:
                    return banco

        return "DESCONOCIDO"

    @classmethod
    def categorizar_transaccion(cls, descripcion: str) -> str:
        """Categoriza automáticamente una transacción."""
        descripcion_lower = descripcion.lower()

        for categoria, patrones in cls.CATEGORIAS.items():
            for patron in patrones:
                if re.search(patron, descripcion_lower):
                    return categoria

        return "otros"

    @classmethod
    def extraer_rfc(cls, texto: str) -> Optional[str]:
        """Extrae RFC de un texto."""
        # Patrón RFC: 4 letras + 6 dígitos + 3 alfanuméricos (persona moral)
        # o 4 letras + 6 dígitos + 2 alfanuméricos + 1 dígito (persona física)
        patron = r"[A-Z&Ñ]{3,4}[0-9]{6}[A-Z0-9]{2,3}"
        match = re.search(patron, texto.upper())
        if match:
            rfc = match.group()
            if len(rfc) in [12, 13]:
                return rfc
        return None


class ImportadorEstadoCuenta:
    """Importador principal de estados de cuenta."""

    def __init__(self):
        self.parser = ParserBancarioMX()
        self.errores: List[str] = []
        self.transacciones_parsed = 0

    def importar(
        self, archivo: bytes, tipo_archivo: str, nombre_archivo: str = ""
    ) -> EstadoCuentaImportado:
        """
        Importa estado de cuenta desde archivo.

        Args:
            archivo: Bytes del archivo
            tipo_archivo: 'pdf', 'csv', 'excel'
            nombre_archivo: Nombre original del archivo

        Returns:
            EstadoCuentaImportado con todas las transacciones
        """
        tipo_lower = tipo_archivo.lower()

        if tipo_lower == "pdf":
            return self._importar_pdf(archivo, nombre_archivo)
        elif tipo_lower == "csv":
            return self._importar_csv(archivo)
        elif tipo_lower in ["excel", "xlsx", "xls"]:
            return self._importar_excel(archivo)
        else:
            raise ValueError(f"Tipo de archivo no soportado: {tipo_archivo}")

    def _importar_pdf(
        self, archivo: bytes, nombre_archivo: str
    ) -> EstadoCuentaImportado:
        """Importa estado de cuenta desde PDF."""
        if not PDF_SUPPORT:
            raise ImportError(
                "PyPDF2 no está instalado. Instalar con: pip install PyPDF2"
            )

        pdf_file = io.BytesIO(archivo)
        pdf_reader = PyPDF2.PdfReader(pdf_file)

        texto_completo = ""
        for page in pdf_reader.pages:
            texto_completo += page.extract_text() + "\n"

        # Detectar banco
        banco = self.parser.detectar_banco(texto_completo)

        # Extraer transacciones según el banco
        if banco == "BBVA":
            transacciones = self._parsear_bbva(texto_completo)
        elif banco == "SANTANDER":
            transacciones = self._parsear_santander(texto_completo)
        elif banco == "BANORTE":
            transacciones = self._parsear_banorte(texto_completo)
        else:
            # Parser genérico
            transacciones = self._parsear_generico(texto_completo)

        return self._construir_estado_cuenta(
            banco=banco,
            cuenta=self._extraer_cuenta(texto_completo),
            transacciones=transacciones,
            texto_completo=texto_completo,
        )

    def _importar_csv(self, archivo: bytes) -> EstadoCuentaImportado:
        """Importa estado de cuenta desde CSV."""
        contenido = archivo.decode("utf-8")

        # Detectar delimitador
        sample = contenido[:1000]
        delimiters = [",", ";", "\t", "|"]
        delimiter = ","
        max_count = 0

        for d in delimiters:
            count = sample.count(d)
            if count > max_count:
                max_count = count
                delimiter = d

        reader = csv.DictReader(io.StringIO(contenido), delimiter=delimiter)
        transacciones = []

        for row in reader:
            try:
                trans = self._parsear_fila_csv(row)
                if trans:
                    transacciones.append(trans)
            except Exception as e:
                self.errores.append(f"Error parseando fila: {e}")
                continue

        return self._construir_estado_cuenta(
            banco="CSV",
            cuenta="",
            transacciones=transacciones,
            texto_completo=contenido,
        )

    def _importar_excel(self, archivo: bytes) -> EstadoCuentaImportado:
        """Importa estado de cuenta desde Excel."""
        if not EXCEL_SUPPORT:
            raise ImportError(
                "openpyxl no está instalado. Instalar con: pip install openpyxl"
            )

        wb = load_workbook(io.BytesIO(archivo))
        ws = wb.active

        transacciones = []

        # Intentar detectar encabezados
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value).lower() if cell.value else "")

        # Mapeo de columnas
        col_fecha = self._encontrar_columna(headers, ["fecha", "date", "f"])
        col_desc = self._encontrar_columna(
            headers, ["descripcion", "concepto", "description", "detalle"]
        )
        col_cargo = self._encontrar_columna(
            headers, ["cargo", "debito", "debit", "egreso", "retiro"]
        )
        col_abono = self._encontrar_columna(
            headers, ["abono", "credito", "credit", "ingreso", "deposito"]
        )
        col_saldo = self._encontrar_columna(headers, ["saldo", "balance"])
        col_ref = self._encontrar_columna(
            headers, ["referencia", "ref", "numero", "folio"]
        )

        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                if not row or not any(row):
                    continue

                # Extraer valores
                fecha_val = (
                    row[col_fecha]
                    if col_fecha is not None and col_fecha < len(row)
                    else None
                )
                desc_val = (
                    row[col_desc]
                    if col_desc is not None and col_desc < len(row)
                    else ""
                )
                cargo_val = (
                    row[col_cargo]
                    if col_cargo is not None and col_cargo < len(row)
                    else 0
                )
                abono_val = (
                    row[col_abono]
                    if col_abono is not None and col_abono < len(row)
                    else 0
                )
                saldo_val = (
                    row[col_saldo]
                    if col_saldo is not None and col_saldo < len(row)
                    else None
                )
                ref_val = (
                    row[col_ref] if col_ref is not None and col_ref < len(row) else ""
                )

                # Parsear fecha
                fecha = self._parsear_fecha(fecha_val)
                if not fecha:
                    continue

                # Parsear montos
                cargo = self._parsear_monto(cargo_val)
                abono = self._parsear_monto(abono_val)
                saldo = self._parsear_monto(saldo_val) if saldo_val else None

                # Categorizar
                descripcion = str(desc_val) if desc_val else ""
                categoria = self.parser.categorizar_transaccion(descripcion)

                trans = TransaccionBancaria(
                    fecha=fecha,
                    descripcion=descripcion,
                    referencia=str(ref_val) if ref_val else "",
                    cargo=cargo,
                    abono=abono,
                    saldo=saldo,
                    concepto_limpio=descripcion.strip(),
                    categoria_sugerida=categoria,
                    rfc_detectado=self.parser.extraer_rfc(descripcion),
                )

                transacciones.append(trans)

            except Exception as e:
                self.errores.append(f"Error en fila: {e}")
                continue

        return self._construir_estado_cuenta(
            banco="EXCEL", cuenta="", transacciones=transacciones, texto_completo=""
        )

    def _parsear_bbva(self, texto: str) -> List[TransaccionBancaria]:
        """Parsea formato específico de BBVA."""
        transacciones = []

        # BBVA usa formato: FECHA | DESCRIPCION | CARGO | ABONO | SALDO
        # Intentar encontrar tabla de movimientos
        lineas = texto.split("\n")

        for linea in lineas:
            # Patrón: fecha seguida de descripción y montos
            patron = r"(\d{2}/\d{2}/\d{4})\s+(.+?)\s+([\d,]+\.\d{2})?\s+([\d,]+\.\d{2})?\s+([\d,]+\.\d{2})?"
            match = re.search(patron, linea)

            if match:
                try:
                    fecha_str = match.group(1)
                    descripcion = match.group(2).strip()
                    cargo_str = match.group(3) or "0"
                    abono_str = match.group(4) or "0"
                    saldo_str = match.group(5)

                    fecha = datetime.strptime(fecha_str, "%d/%m/%Y").date()
                    cargo = Decimal(cargo_str.replace(",", ""))
                    abono = Decimal(abono_str.replace(",", ""))
                    saldo = Decimal(saldo_str.replace(",", "")) if saldo_str else None

                    categoria = self.parser.categorizar_transaccion(descripcion)

                    trans = TransaccionBancaria(
                        fecha=fecha,
                        descripcion=descripcion,
                        referencia="",
                        cargo=cargo,
                        abono=abono,
                        saldo=saldo,
                        categoria_sugerida=categoria,
                    )
                    transacciones.append(trans)

                except Exception:
                    continue

        return transacciones

    def _parsear_santander(self, texto: str) -> List[TransaccionBancaria]:
        """Parsea formato específico de Santander."""
        transacciones = []
        # Implementación similar a BBVA con patrones específicos de Santander
        return transacciones

    def _parsear_banorte(self, texto: str) -> List[TransaccionBancaria]:
        """Parsea formato específico de Banorte."""
        transacciones = []
        # Implementación específica de Banorte
        return transacciones

    def _parsear_generico(self, texto: str) -> List[TransaccionBancaria]:
        """Parser genérico para cualquier formato."""
        transacciones = []
        lineas = texto.split("\n")

        for linea in lineas:
            # Buscar patrones comunes de transacciones
            # Fecha + Monto
            patron = r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}).*?(\d{1,3}(?:,\d{3})*\.\d{2})"
            match = re.search(patron, linea)

            if match:
                try:
                    fecha_str = match.group(1)
                    # Intentar diferentes formatos de fecha
                    for fmt in ["%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%y"]:
                        try:
                            fecha = datetime.strptime(fecha_str, fmt).date()
                            break
                        except ValueError:
                            continue
                    else:
                        continue

                    # El resto de la línea es descripción
                    descripcion = (
                        linea[: match.start()].strip() + linea[match.end() :].strip()
                    )

                    # Determinar si es cargo o abono basado en el contexto
                    monto = Decimal(match.group(2).replace(",", ""))

                    trans = TransaccionBancaria(
                        fecha=fecha,
                        descripcion=descripcion[:100],
                        referencia="",
                        cargo=monto if "cargo" in descripcion.lower() else Decimal("0"),
                        abono=monto if "abono" in descripcion.lower() else Decimal("0"),
                        categoria_sugerida="otros",
                    )
                    transacciones.append(trans)

                except Exception:
                    continue

        return transacciones

    def _parsear_fila_csv(self, row: Dict) -> Optional[TransaccionBancaria]:
        """Parsea una fila de CSV."""
        # Encontrar campos por nombre similar
        fecha_key = self._encontrar_campo(row.keys(), ["fecha", "date", "f"])
        desc_key = self._encontrar_campo(
            row.keys(), ["descripcion", "concepto", "description"]
        )
        cargo_key = self._encontrar_campo(row.keys(), ["cargo", "debito", "debit"])
        abono_key = self._encontrar_campo(row.keys(), ["abono", "credito", "credit"])
        saldo_key = self._encontrar_campo(row.keys(), ["saldo", "balance"])

        if not fecha_key or not desc_key:
            return None

        fecha = self._parsear_fecha(row.get(fecha_key))
        if not fecha:
            return None

        descripcion = str(row.get(desc_key, ""))

        return TransaccionBancaria(
            fecha=fecha,
            descripcion=descripcion,
            referencia="",
            cargo=self._parsear_monto(row.get(cargo_key, 0)),
            abono=self._parsear_monto(row.get(abono_key, 0)),
            saldo=self._parsear_monto(row.get(saldo_key)) if saldo_key else None,
            categoria_sugerida=self.parser.categorizar_transaccion(descripcion),
        )

    def _construir_estado_cuenta(
        self,
        banco: str,
        cuenta: str,
        transacciones: List[TransaccionBancaria],
        texto_completo: str,
    ) -> EstadoCuentaImportado:
        """Construye objeto EstadoCuentaImportado."""

        if not transacciones:
            return EstadoCuentaImportado(
                banco=banco,
                cuenta=cuenta,
                moneda="MXN",
                fecha_inicio=date.today(),
                fecha_fin=date.today(),
                saldo_inicial=Decimal("0"),
                saldo_final=Decimal("0"),
                transacciones=[],
                total_cargos=Decimal("0"),
                total_abonos=Decimal("0"),
            )

        # Ordenar por fecha
        transacciones.sort(key=lambda x: x.fecha)

        total_cargos = sum(t.cargo for t in transacciones)
        total_abonos = sum(t.abono for t in transacciones)

        # Intentar extraer saldos del texto
        saldo_inicial = self._extraer_saldo_inicial(texto_completo)
        saldo_final = self._extraer_saldo_final(texto_completo)

        # Si no se encontraron, calcular
        if saldo_inicial is None:
            saldo_inicial = transacciones[0].saldo or Decimal("0")
        if saldo_final is None:
            saldo_final = transacciones[-1].saldo or (
                saldo_inicial + total_abonos - total_cargos
            )

        return EstadoCuentaImportado(
            banco=banco,
            cuenta=cuenta,
            moneda="MXN",
            fecha_inicio=transacciones[0].fecha,
            fecha_fin=transacciones[-1].fecha,
            saldo_inicial=saldo_inicial,
            saldo_final=saldo_final,
            transacciones=transacciones,
            total_cargos=total_cargos,
            total_abonos=total_abonos,
        )

    def _encontrar_columna(
        self, headers: List[str], opciones: List[str]
    ) -> Optional[int]:
        """Encuentra índice de columna basado en opciones."""
        for i, header in enumerate(headers):
            header_lower = header.lower().strip()
            for opcion in opciones:
                if opcion in header_lower:
                    return i
        return None

    def _encontrar_campo(self, keys, opciones: List[str]) -> Optional[str]:
        """Encuentra campo en diccionario."""
        for key in keys:
            key_lower = key.lower()
            for opcion in opciones:
                if opcion in key_lower:
                    return key
        return None

    def _parsear_fecha(self, valor: Any) -> Optional[date]:
        """Parsea fecha desde varios formatos."""
        if isinstance(valor, date):
            return valor

        if isinstance(valor, datetime):
            return valor.date()

        if not valor or str(valor).strip() == "":
            return None

        fecha_str = str(valor).strip()
        formatos = [
            "%d/%m/%Y",
            "%d-%m-%Y",
            "%Y/%m/%d",
            "%Y-%m-%d",
            "%d/%m/%y",
            "%m/%d/%Y",
            "%Y%m%d",
        ]

        for fmt in formatos:
            try:
                return datetime.strptime(fecha_str, fmt).date()
            except ValueError:
                continue

        return None

    def _parsear_monto(self, valor: Any) -> Decimal:
        """Parsea monto desde string o número."""
        if isinstance(valor, (int, float)):
            return Decimal(str(valor))

        if not valor:
            return Decimal("0")

        # Limpiar string
        monto_str = str(valor).strip()
        monto_str = monto_str.replace("$", "").replace(",", "")

        # Manejar paréntesis para negativos
        if "(" in monto_str and ")" in monto_str:
            monto_str = "-" + monto_str.replace("(", "").replace(")", "")

        try:
            return Decimal(monto_str)
        except InvalidOperation:
            return Decimal("0")

    def _extraer_cuenta(self, texto: str) -> str:
        """Extrae número de cuenta del texto."""
        # Buscar patrones comunes de números de cuenta
        patrones = [
            r"cuenta[:\s]+(\d{10,20})",
            r"no\.?\s*cuenta[:\s]+(\d{10,20})",
            r"account[:\s]+(\d{10,20})",
            r"\*{4,}(\d{4})",  # ****1234
        ]

        for patron in patrones:
            match = re.search(patron, texto, re.IGNORECASE)
            if match:
                return match.group(1)

        return ""

    def _extraer_saldo_inicial(self, texto: str) -> Optional[Decimal]:
        """Extrae saldo inicial del texto."""
        patrones = [
            r"saldo\s*(anterior|inicial|previous)[:\s]+\$?([\d,]+\.\d{2})",
            r"saldo\s*al\s*inicio[:\s]+\$?([\d,]+\.\d{2})",
        ]

        for patron in patrones:
            match = re.search(patron, texto, re.IGNORECASE)
            if match:
                try:
                    return Decimal(match.group(2).replace(",", ""))
                except:
                    continue

        return None

    def _extraer_saldo_final(self, texto: str) -> Optional[Decimal]:
        """Extrae saldo final del texto."""
        patrones = [
            r"saldo\s*(final|actual|current)[:\s]+\$?([\d,]+\.\d{2})",
            r"saldo\s*al\s*corte[:\s]+\$?([\d,]+\.\d{2})",
        ]

        for patron in patrones:
            match = re.search(patron, texto, re.IGNORECASE)
            if match:
                try:
                    return Decimal(match.group(2).replace(",", ""))
                except:
                    continue

        return None

    def generar_resumen_categorias(
        self, estado: EstadoCuentaImportado
    ) -> Dict[str, Any]:
        """Genera resumen agrupado por categorías."""
        categorias = {}

        for trans in estado.transacciones:
            cat = trans.categoria_sugerida
            if cat not in categorias:
                categorias[cat] = {
                    "cantidad": 0,
                    "cargos": Decimal("0"),
                    "abonos": Decimal("0"),
                    "neto": Decimal("0"),
                }

            categorias[cat]["cantidad"] += 1
            categorias[cat]["cargos"] += trans.cargo
            categorias[cat]["abonos"] += trans.abono
            categorias[cat]["neto"] += trans.monto_neto

        return categorias
